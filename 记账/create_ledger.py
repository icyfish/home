#!/usr/bin/env python3
"""生成 momo & star 记账表"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── 颜色和样式定义 ──────────────────────────────────────────────
HEADER_FONT = Font(name="PingFang SC", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MOMO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
STAR_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_range(ws, start_row, end_row, cols, fill=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if c != cols else LEFT  # 最后一列左对齐（备注）
            if fill:
                cell.fill = fill


# ── Sheet 1: 消费类目 ──────────────────────────────────────────
ws_cat = wb.active
ws_cat.title = "消费类目"
ws_cat.sheet_properties.tabColor = "4472C4"

cat_headers = ["序号", "类目名称", "说明"]
ws_cat.append(cat_headers)
style_header(ws_cat, 1, len(cat_headers))

default_categories = [
    ("餐饮", "一日三餐、外卖、聚餐等"),
    ("交通", "公交、地铁、打车、加油等"),
    ("购物", "日用品、服饰、电子产品等"),
    ("住房", "房租、水电燃气、物业费等"),
    ("娱乐", "电影、游戏、旅行、运动等"),
    ("医疗", "看病、药品、体检等"),
    ("教育", "课程、书籍、培训等"),
    ("人情", "礼金、红包、请客等"),
    ("宠物", "宠物食品、医疗、用品等"),
    ("其他", "未分类支出"),
]

for i, (name, desc) in enumerate(default_categories, 1):
    ws_cat.append([i, name, desc])

style_range(ws_cat, 2, len(default_categories) + 1, len(cat_headers), SUB_FILL)

# 预留空行供用户添加
for i in range(len(default_categories) + 1, len(default_categories) + 21):
    ws_cat.append([i, "", ""])
    style_range(ws_cat, i + 1, i + 1, len(cat_headers))

ws_cat.column_dimensions["A"].width = 8
ws_cat.column_dimensions["B"].width = 16
ws_cat.column_dimensions["C"].width = 35


# ── Sheet 2: 预算来源 ──────────────────────────────────────────
ws_src = wb.create_sheet("预算来源")
ws_src.sheet_properties.tabColor = "70AD47"

src_headers = ["序号", "来源名称", "所属人", "说明"]
ws_src.append(src_headers)
style_header(ws_src, 1, len(src_headers))

default_sources = [
    ("工资", "momo", "每月固定工资收入"),
    ("工资", "star", "每月固定工资收入"),
    ("兼职", "momo", "兼职/副业收入"),
    ("兼职", "star", "兼职/副业收入"),
    ("投资收益", "共同", "理财、基金、股票等收益"),
    ("红包", "共同", "节日红包、礼金等"),
    ("报销", "共同", "公司报销款"),
    ("其他", "共同", "其他收入来源"),
]

for i, (name, owner, desc) in enumerate(default_sources, 1):
    ws_src.append([i, name, owner, desc])

style_range(ws_src, 2, len(default_sources) + 1, len(src_headers), SUB_FILL)

# 预留空行
for i in range(len(default_sources) + 1, len(default_sources) + 21):
    ws_src.append([i, "", "", ""])
    style_range(ws_src, i + 1, i + 1, len(src_headers))

ws_src.column_dimensions["A"].width = 8
ws_src.column_dimensions["B"].width = 16
ws_src.column_dimensions["C"].width = 12
ws_src.column_dimensions["D"].width = 30


# ── Sheet 3: 记账明细 ──────────────────────────────────────────
ws_rec = wb.create_sheet("记账明细")
ws_rec.sheet_properties.tabColor = "ED7D31"
wb.move_sheet("记账明细", offset=-2)  # 移到第一个位置

rec_headers = ["日期", "记账人", "收/支", "类目", "来源", "金额", "备注"]
ws_rec.append(rec_headers)
style_header(ws_rec, 1, len(rec_headers))

# 数据验证 - 记账人下拉
dv_user = DataValidation(type="list", formula1='"momo,star"', allow_blank=True)
dv_user.error = "请选择 momo 或 star"
dv_user.errorTitle = "无效输入"
ws_rec.add_data_validation(dv_user)
dv_user.add("B2:B1000")

# 数据验证 - 收/支下拉
dv_type = DataValidation(type="list", formula1='"支出,收入"', allow_blank=True)
ws_rec.add_data_validation(dv_type)
dv_type.add("C2:C1000")

# 数据验证 - 类目下拉（引用消费类目表）
cat_range = f"消费类目!$B$2:$B${len(default_categories) + 20}"
dv_cat = DataValidation(type="list", formula1=f"={cat_range}", allow_blank=True)
dv_cat.error = "请从消费类目中选择，或先到【消费类目】表添加新类目"
ws_rec.add_data_validation(dv_cat)
dv_cat.add("D2:D1000")

# 数据验证 - 来源下拉（引用预算来源表）
src_range = f"预算来源!$B$2:$B${len(default_sources) + 20}"
dv_src = DataValidation(type="list", formula1=f"={src_range}", allow_blank=True)
dv_src.error = "请从预算来源中选择，或先到【预算来源】表添加新来源"
ws_rec.add_data_validation(dv_src)
dv_src.add("E2:E1000")

# 数据验证 - 日期格式
dv_date = DataValidation(type="date", allow_blank=True)
dv_date.error = "请输入有效日期，格式如 2026-03-01"
ws_rec.add_data_validation(dv_date)
dv_date.add("A2:A1000")

# 设置金额列为数字格式
for r in range(2, 1001):
    ws_rec.cell(row=r, column=6).number_format = '#,##0.00'

# 设置日期列格式
for r in range(2, 1001):
    ws_rec.cell(row=r, column=1).number_format = 'YYYY-MM-DD'

ws_rec.column_dimensions["A"].width = 14
ws_rec.column_dimensions["B"].width = 10
ws_rec.column_dimensions["C"].width = 8
ws_rec.column_dimensions["D"].width = 12
ws_rec.column_dimensions["E"].width = 14
ws_rec.column_dimensions["F"].width = 14
ws_rec.column_dimensions["G"].width = 30

# 冻结首行
ws_rec.freeze_panes = "A2"


# ── Sheet 4: 月度汇总 ──────────────────────────────────────────
ws_sum = wb.create_sheet("月度汇总")
ws_sum.sheet_properties.tabColor = "7030A0"

sum_headers = ["月份", "记账人", "总支出", "总收入", "结余"]
ws_sum.append(sum_headers)
style_header(ws_sum, 1, len(sum_headers))

# 添加示例公式说明行
ws_sum.cell(row=2, column=1, value="使用说明：").font = Font(bold=True, color="FF0000")
ws_sum.merge_cells("A2:E2")
ws_sum.cell(row=3, column=1, value="在下方手动填写月份和记账人，金额可用 SUMPRODUCT 公式自动计算").font = Font(
    italic=True, color="808080"
)
ws_sum.merge_cells("A3:E3")

# momo 示例行
r = 5
ws_sum.cell(row=r, column=1, value="2026-03")
ws_sum.cell(row=r, column=2, value="momo")
ws_sum.cell(row=r, column=3).value = '=SUMPRODUCT((记账明细!B$2:B$1000="momo")*(记账明细!C$2:C$1000="支出")*(TEXT(记账明细!A$2:A$1000,"YYYY-MM")=A5)*记账明细!F$2:F$1000)'
ws_sum.cell(row=r, column=4).value = '=SUMPRODUCT((记账明细!B$2:B$1000="momo")*(记账明细!C$2:C$1000="收入")*(TEXT(记账明细!A$2:A$1000,"YYYY-MM")=A5)*记账明细!F$2:F$1000)'
ws_sum.cell(row=r, column=5).value = "=D5-C5"

# star 示例行
r = 6
ws_sum.cell(row=r, column=1, value="2026-03")
ws_sum.cell(row=r, column=2, value="star")
ws_sum.cell(row=r, column=3).value = '=SUMPRODUCT((记账明细!B$2:B$1000="star")*(记账明细!C$2:C$1000="支出")*(TEXT(记账明细!A$2:A$1000,"YYYY-MM")=A6)*记账明细!F$2:F$1000)'
ws_sum.cell(row=r, column=4).value = '=SUMPRODUCT((记账明细!B$2:B$1000="star")*(记账明细!C$2:C$1000="收入")*(TEXT(记账明细!A$2:A$1000,"YYYY-MM")=A6)*记账明细!F$2:F$1000)'
ws_sum.cell(row=r, column=5).value = "=D6-C6"

for row in range(5, 7):
    for c in range(3, 6):
        ws_sum.cell(row=row, column=c).number_format = '#,##0.00'
    style_range(ws_sum, row, row, len(sum_headers))

ws_sum.column_dimensions["A"].width = 14
ws_sum.column_dimensions["B"].width = 10
ws_sum.column_dimensions["C"].width = 14
ws_sum.column_dimensions["D"].width = 14
ws_sum.column_dimensions["E"].width = 14

ws_sum.freeze_panes = "A2"


# ── 保存 ────────────────────────────────────────────────────────
output = "记账表_momo_star.xlsx"
wb.save(output)
print(f"✅ 记账表已生成: {output}")
